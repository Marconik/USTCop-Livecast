#!/usr/bin/env python3
"""Generate a tournament bracket image from a group schedule workbook."""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:  # pragma: no cover - user-facing dependency guard
    missing = exc.name or "required package"
    raise SystemExit(
        f"Missing Python package: {missing}. Install dependencies with:\n"
        f"  python -m pip install pillow openpyxl"
    ) from exc


BASE_DIR = Path(__file__).resolve().parent


@dataclass
class Player:
    no: str = ""
    name: str = ""

    @property
    def has_value(self) -> bool:
        return bool(self.no or self.name)


@dataclass
class Match:
    players: list[Player]
    winner_slot: int | None = None


@dataclass(frozen=True)
class SlotLayout:
    round_no: int
    slot_index: int
    anchor: str
    x: int
    y: int


# Coordinates are for raw.png at 5011x2818. The x/y values are anchors:
# left/right center for side-aligned boxes, or left center for center boxes.
LAYOUT: tuple[SlotLayout, ...] = (
    SlotLayout(1, 0, "right_center", 801, 833),
    SlotLayout(1, 1, "right_center", 801, 1387),
    SlotLayout(4, 0, "right_center", 801, 1955),
    SlotLayout(4, 1, "right_center", 801, 2484),
    SlotLayout(5, 0, "left_center", 1020, 1112),
    SlotLayout(5, 1, "left_center", 1020, 2219),
    SlotLayout(7, 0, "left_center", 1585, 1709),
    SlotLayout(7, 1, "left_center", 2646, 1709),
    SlotLayout(6, 0, "right_center", 3800, 1123),
    SlotLayout(6, 1, "right_center", 3800, 2210),
    SlotLayout(2, 0, "left_center", 4118, 833),
    SlotLayout(2, 1, "left_center", 4118, 1387),
    SlotLayout(3, 0, "left_center", 4118, 1955),
    SlotLayout(3, 1, "left_center", 4118, 2484),
)

GROUP_LABEL_CENTER = (2455, 356)

FEEDERS: dict[int, tuple[int, int]] = {
    5: (1, 4),
    6: (2, 3),
    7: (5, 6),
}


def clean_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def is_winner_marker(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return value == 1
    return clean_cell(value) == "1"


def group_label_from_path(xlsx_path: Path) -> str:
    stem = xlsx_path.stem
    if stem.lower().startswith("group") and len(stem) > 5:
        return stem[5:].upper()
    return stem.upper()


def resolve_group_file(group: str | None, xlsx: str | None) -> Path:
    if xlsx:
        return Path(xlsx).resolve()
    if not group:
        raise SystemExit("Please provide --group A/B/C/D/E or --xlsx path.")

    group_name = group.strip()
    if len(group_name) == 1:
        group_name = group_name.upper()
    if group_name.upper().startswith("GROUP"):
        filename = f"{group_name}.xlsx"
    else:
        filename = f"Group{group_name}.xlsx"
    return (BASE_DIR / filename).resolve()


def header_map(headers: list[Any]) -> dict[str, int]:
    names = {clean_cell(value): idx for idx, value in enumerate(headers)}
    fallback = {
        "回合": 0,
        "选手序号": 1,
        "选手名": 2,
        "获胜": 7,
    }
    for name, idx in fallback.items():
        names.setdefault(name, idx)
    return names


def read_matches(xlsx_path: Path, sheet_name: str | None = None) -> dict[int, Match]:
    if not xlsx_path.exists():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Workbook is empty: {xlsx_path}")

    headers = header_map(list(rows[0]))
    grouped: dict[int, list[tuple[Player, bool]]] = {}

    for row in rows[1:]:
        round_text = clean_cell(row[headers["回合"]] if len(row) > headers["回合"] else "")
        if not round_text:
            continue
        try:
            round_no = int(float(round_text))
        except ValueError:
            continue

        no = clean_cell(row[headers["选手序号"]] if len(row) > headers["选手序号"] else "")
        name = clean_cell(row[headers["选手名"]] if len(row) > headers["选手名"] else "")
        won = is_winner_marker(row[headers["获胜"]] if len(row) > headers["获胜"] else None)
        grouped.setdefault(round_no, []).append((Player(no=no, name=name), won))

    matches: dict[int, Match] = {}
    for round_no in range(1, 8):
        entries = grouped.get(round_no, [])
        players = [entry[0] for entry in entries[:2]]
        while len(players) < 2:
            players.append(Player())

        winner_slot = None
        for idx, (_, won) in enumerate(entries[:2]):
            if won:
                winner_slot = idx
                break
        matches[round_no] = Match(players=players, winner_slot=winner_slot)

    return matches


def winner_of(match: Match) -> Player:
    if match.winner_slot is None:
        return Player()
    return match.players[match.winner_slot]


def complete_advancement(matches: dict[int, Match]) -> dict[int, Match]:
    completed = {round_no: Match(players=list(match.players), winner_slot=match.winner_slot) for round_no, match in matches.items()}

    for round_no in (5, 6, 7):
        feeder_a, feeder_b = FEEDERS[round_no]
        feeder_players = [winner_of(completed[feeder_a]), winner_of(completed[feeder_b])]
        current = completed[round_no]
        players = []
        for idx, sheet_player in enumerate(current.players):
            players.append(sheet_player if sheet_player.has_value else feeder_players[idx])
        completed[round_no] = Match(players=players, winner_slot=current.winner_slot)

    return completed


def load_font(candidates: list[Path], size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in candidates:
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size)
            except OSError:
                continue
    return ImageFont.load_default()


def fit_font(
    draw: ImageDraw.ImageDraw,
    text: str,
    candidates: list[Path],
    max_size: int,
    min_size: int,
    max_width: int,
) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for size in range(max_size, min_size - 1, -2):
        font = load_font(candidates, size)
        bbox = draw.textbbox((0, 0), text, font=font, stroke_width=0)
        if bbox[2] - bbox[0] <= max_width:
            return font
    return load_font(candidates, min_size)


def anchored_position(layout: SlotLayout, box: Image.Image) -> tuple[int, int]:
    width, height = box.size
    if layout.anchor == "right_center":
        return layout.x - width, layout.y - height // 2
    if layout.anchor == "left_center":
        return layout.x, layout.y - height // 2
    if layout.anchor == "center":
        return layout.x - width // 2, layout.y - height // 2
    raise ValueError(f"Unsupported anchor: {layout.anchor}")


def draw_centered_text(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: tuple[int, int, int],
    stroke_fill: tuple[int, int, int] | None = None,
    stroke_width: int = 0,
) -> None:
    bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke_width)
    x = xy[0] - (bbox[2] - bbox[0]) // 2 - bbox[0]
    y = xy[1] - (bbox[3] - bbox[1]) // 2 - bbox[1]
    draw.text((x, y), text, font=font, fill=fill, stroke_fill=stroke_fill, stroke_width=stroke_width)


def draw_centered_squeezed_text(
    canvas: Image.Image,
    xy: tuple[int, int],
    text: str,
    font: ImageFont.FreeTypeFont | ImageFont.ImageFont,
    fill: tuple[int, int, int],
    max_width: int,
) -> None:
    draw = ImageDraw.Draw(canvas)
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    if text_width <= 0 or text_height <= 0:
        return

    layer = Image.new("RGBA", (text_width, text_height), (255, 255, 255, 0))
    layer_draw = ImageDraw.Draw(layer)
    layer_draw.text((-bbox[0], -bbox[1]), text, font=font, fill=fill)

    if text_width > max_width:
        layer = layer.resize((max_width, text_height), Image.Resampling.LANCZOS)

    left = xy[0] - layer.width // 2
    top = xy[1] - layer.height // 2
    canvas.alpha_composite(layer, (left, top))


def draw_player_box(
    canvas: Image.Image,
    box: Image.Image,
    top_left: tuple[int, int],
    player: Player,
    font_paths: dict[str, list[Path]],
) -> None:
    canvas.alpha_composite(box, top_left)
    draw = ImageDraw.Draw(canvas)

    x, y = top_left
    width, height = box.size
    no_text = f"No.{player.no}" if player.no else "No."

    no_font = fit_font(draw, no_text, font_paths["no"], 88, 48, int(width * 0.68))
    name_font = load_font(font_paths["name"], 66) if player.name else None

    draw_centered_text(
        draw,
        (x + width // 2, y + int(height * 0.34)),
        no_text,
        no_font,
        fill=(11, 51, 91),
        stroke_fill=(226, 244, 248),
        stroke_width=3,
    )

    if player.name:
        draw_centered_squeezed_text(
            canvas,
            (x + width // 2, y + int(height * 0.66)),
            player.name,
            name_font,
            fill=(0, 0, 0),
            max_width=int(width * 0.72),
        )


def draw_group_label(
    canvas: Image.Image,
    label: str,
    font_paths: dict[str, list[Path]],
    scale_x: float,
    scale_y: float,
) -> None:
    if not label:
        return

    display_label = label.strip().upper()[:1]
    if not display_label:
        return

    draw = ImageDraw.Draw(canvas)
    font = load_font(font_paths["group"], round(230 * min(scale_x, scale_y)))
    center = (round(GROUP_LABEL_CENTER[0] * scale_x), round(GROUP_LABEL_CENTER[1] * scale_y))
    draw_centered_text(
        draw,
        center,
        display_label,
        font,
        fill=(4, 42, 108),
        stroke_fill=(255, 255, 255),
        stroke_width=2,
    )


def generate_bracket(
    matches: dict[int, Match],
    background_path: Path,
    normal_box_path: Path,
    winner_box_path: Path,
    output_path: Path,
    group_label: str = "",
) -> None:
    background = Image.open(background_path).convert("RGBA")
    normal_box = Image.open(normal_box_path).convert("RGBA")
    winner_box = Image.open(winner_box_path).convert("RGBA")

    scale_x = background.width / 5011
    scale_y = background.height / 2818

    if (background.width, background.height) != (5011, 2818):
        normal_box = normal_box.resize((round(normal_box.width * scale_x), round(normal_box.height * scale_y)), Image.Resampling.LANCZOS)
        winner_box = winner_box.resize((round(winner_box.width * scale_x), round(winner_box.height * scale_y)), Image.Resampling.LANCZOS)

    font_paths = {
        "no": [
            BASE_DIR / "image-gen" / "Billiton Gothic.ttf",
            BASE_DIR / "image-gen" / "BankGothic Lt BT Light.ttf",
            Path(r"C:\Windows\Fonts\arialbd.ttf"),
        ],
        "name": [
            BASE_DIR / "image-gen" / "BaseMono-Narrow.ttf",
            Path(r"C:\Windows\Fonts\msyh.ttc"),
            Path(r"C:\Windows\Fonts\simhei.ttf"),
            Path(r"C:\Windows\Fonts\simsun.ttc"),
            Path(r"C:\Windows\Fonts\arial.ttf"),
        ],
        "group": [
            Path(r"C:\Windows\Fonts\arialbd.ttf"),
            BASE_DIR / "image-gen" / "Billiton Gothic.ttf",
            BASE_DIR / "image-gen" / "BankGothic Lt BT Light.ttf",
        ],
    }

    draw_group_label(background, group_label, font_paths, scale_x, scale_y)

    for slot in LAYOUT:
        match = matches[slot.round_no]
        is_winner = match.winner_slot == slot.slot_index
        box = winner_box if is_winner else normal_box
        layout = SlotLayout(
            slot.round_no,
            slot.slot_index,
            slot.anchor,
            round(slot.x * scale_x),
            round(slot.y * scale_y),
        )
        top_left = anchored_position(layout, box)
        draw_player_box(background, box, top_left, match.players[slot.slot_index], font_paths)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    background.convert("RGB").save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a tournament bracket image.")
    parser.add_argument("-g", "--group", help="Group name, for example A, B, C, D or E.")
    parser.add_argument("-x", "--xlsx", help="Schedule workbook path. Overrides --group.")
    parser.add_argument("-s", "--sheet", help="Worksheet name. Defaults to the first sheet.")
    parser.add_argument("-o", "--output", help="Output PNG path. Defaults to outputs/bracket_<group>.png.")
    parser.add_argument("--group-label", help="Letter shown in the top circle. Defaults to --group or the workbook name.")
    parser.add_argument("--background", default=str(BASE_DIR / "raw.png"), help="Background image path.")
    parser.add_argument("--box1", default=str(BASE_DIR / "box1.png"), help="Normal player box image path.")
    parser.add_argument("--box2", default=str(BASE_DIR / "box2.png"), help="Winner player box image path.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    xlsx_path = resolve_group_file(args.group, args.xlsx)
    matches = complete_advancement(read_matches(xlsx_path, args.sheet))
    group_label = args.group_label or args.group or group_label_from_path(xlsx_path)

    if args.output:
        output_path = Path(args.output).resolve()
    else:
        label = args.group or xlsx_path.stem.replace("Group", "")
        output_path = (BASE_DIR / "outputs" / f"bracket_{label}.png").resolve()

    generate_bracket(
        matches=matches,
        background_path=Path(args.background).resolve(),
        normal_box_path=Path(args.box1).resolve(),
        winner_box_path=Path(args.box2).resolve(),
        output_path=output_path,
        group_label=group_label,
    )
    print(f"Generated: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
